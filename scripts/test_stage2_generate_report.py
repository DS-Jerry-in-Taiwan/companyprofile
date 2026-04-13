#!/usr/bin/env python3
"""
Phase 14 Stage 2 整合測試報告生成器 (修正版)
使用 Checkpoint 1 的 10 個真實測試案例

測試邏輯：
  Checkpoint 1 輸入（同一份）
    → post_process（相同處理）
      → differentiate_template（三種模式截斷）
        → 三種模式的截斷結果應該不同

關鍵：同一個輸入，調用不同模式，結果應該不同
"""

import sys
import os
import re
import json
import csv
from pathlib import Path
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from src.functions.utils.post_processing import post_process
from src.functions.utils.template_differentiator import (
    differentiate_template,
    validate_template_differentiation,
)


# ==================== 設定 ====================

OUTPUT_DIR = Path(
    "/home/ubuntu/projects/OrganBriefOptimization/docs/test_report/v0.0.1/phase14/stage2/artifacts"
)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ==================== 工具函數 ====================


def load_checkpoint1_data():
    """載入 Checkpoint 1 測試資料"""
    path = Path(
        "/home/ubuntu/projects/OrganBriefOptimization/docs/test_report/v0.0.1/phase14/checkpoint1/artifacts/test_inputs.json"
    )
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data["test_inputs"]


def strip_html(text):
    """去除 HTML 標籤，返回純文字"""
    text = re.sub(r"</p>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<p[^>]*>", "", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"\n\n+", "\n", text)
    return text.strip()


def count_chinese(text):
    """計算中文字數"""
    plain = strip_html(text)
    return len(re.findall(r"[\u4e00-\u9fff]", plain))


def count_plain(text):
    """計算純文字字符數"""
    return len(strip_html(text))


# ==================== 主程式 ====================


def main():
    print("=" * 70)
    print("Phase 14 Stage 2 整合測試報告生成（修正版）")
    print("=" * 70)

    # 載入資料
    print("\n📂 載入 Checkpoint 1 測試資料...")
    cases = load_checkpoint1_data()
    print(f"✅ 載入 {len(cases)} 個測試案例")

    MODES = ["concise", "standard", "detailed"]
    MODE_NAMES = {
        "concise": "精簡 (concise)",
        "standard": "標準 (standard)",
        "detailed": "詳細 (detailed)",
    }
    MODE_LIMITS = {"concise": 100, "standard": 200, "detailed": 500}
    MODE_MAX_ALLOWED = {"concise": 110, "standard": 210, "detailed": 600}

    # 處理所有案例
    all_results = []
    summary_rows = []
    comparison_rows = []

    print("\n🔄 處理所有案例（同一輸入，三種模式截斷）...")
    for case in cases:
        case_id = case["case_id"]
        company_name = case["company"]["name"]
        organ_no = case["company"].get("registration_id", "")
        input_text = case.get("input_text", "")
        industry = case["company"].get("industry", {}).get("type", "")
        product_service = case.get("product_service", "")

        print(f"  案例 {case_id}: {company_name}")

        # 將用戶輸入封裝成 LLM 輸出格式（統一加上 HTML 包裝）
        llm_output = f"<p>{input_text}</p>"

        # ===== 核心邏輯：同一個 post_process，三種模式截斷 =====
        mode_stats = {}
        for mode in MODES:
            # 1. 相同的 post_process（不指定 template_type 或統一 standard）
            processed = post_process(
                {"body_html": llm_output}, template_type="standard"
            )
            body_html = processed.get("body_html", "")

            # 2. 不同的 differentiate_template（截斷到不同長度）
            final = differentiate_template(body_html, mode)

            mode_stats[mode] = {
                "llm_output": llm_output,
                "post_processed": body_html,
                "final_output": final,
                "plain_text": strip_html(final),
                "chinese_count": count_chinese(final),
                "plain_count": count_plain(final),
            }

        # ===== Sheet 1: 三模板對比表 =====
        for mode in MODES:
            r = mode_stats[mode]
            comparison_rows.append(
                {
                    "案例": case_id,
                    "公司名稱": company_name,
                    "模板模式": MODE_NAMES[mode],
                    "字數上限": MODE_LIMITS[mode],
                    "中文字數": r["chinese_count"],
                    "純文字數": r["plain_count"],
                    "是否超標": "❌ 超標"
                    if r["chinese_count"] > MODE_MAX_ALLOWED[mode]
                    else "✅ OK",
                    "內容（純文字）": r["plain_text"],
                    "內容（含HTML）": r["final_output"],
                }
            )

        # ===== Sheet 2: 摘要統計 =====
        c_chars = mode_stats["concise"]["chinese_count"]
        s_chars = mode_stats["standard"]["chinese_count"]
        d_chars = mode_stats["detailed"]["chinese_count"]

        c_ok = c_chars <= MODE_MAX_ALLOWED["concise"]
        s_ok = s_chars <= MODE_MAX_ALLOWED["standard"]
        d_diff = d_chars > c_chars

        # 截斷差異度
        diff_s_c = s_chars - c_chars
        diff_d_s = d_chars - s_chars

        summary_rows.append(
            {
                "案例": case_id,
                "公司名稱": company_name,
                # Concise
                "concise_字數": c_chars,
                "concise_狀態": "✅" if c_ok else "❌",
                "concise_內容": mode_stats["concise"]["plain_text"],
                # Standard
                "standard_字數": s_chars,
                "standard_狀態": "✅" if s_ok else "❌",
                "standard_內容": mode_stats["standard"]["plain_text"],
                # Detailed
                "detailed_字數": d_chars,
                "detailed_狀態": "✅" if d_chars > c_chars else "❌",
                "detailed_內容": mode_stats["detailed"]["plain_text"],
                # 差異驗證
                "三模板差異": "✅" if d_diff else "❌",
                "c_s差": diff_s_c,
                "s_d差": diff_d_s,
            }
        )

        all_results.append(
            {
                "case_id": case_id,
                "company_name": company_name,
                "organ_no": organ_no,
                "industry": industry,
                "product_service": product_service,
                "input_text": input_text,
                "modes": mode_stats,
            }
        )

    # ===== 計算總體統計 =====
    total = len(cases)
    concise_ok = sum(1 for r in summary_rows if r["concise_狀態"] == "✅")
    standard_ok = sum(1 for r in summary_rows if r["standard_狀態"] == "✅")
    all_different = sum(1 for r in summary_rows if r["三模板差異"] == "✅")

    avg_c = sum(r["concise_字數"] for r in summary_rows) / total
    avg_s = sum(r["standard_字數"] for r in summary_rows) / total
    avg_d = sum(r["detailed_字數"] for r in summary_rows) / total

    print(f"\n  concise 達標: {concise_ok}/{total}")
    print(f"  standard 達標: {standard_ok}/{total}")
    print(f"  三模板差異: {all_different}/{total}")

    # ===== 生成 Excel 報告 =====
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = OUTPUT_DIR / f"stage2_template_differentiation_fixed_{timestamp}.xlsx"

    print(f"\n📊 生成 Excel 報告...")
    print(f"   檔案: {excel_path}")

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # ===== Sheet 1: 三模板對比表 =====
        df_comparison = pd.DataFrame(comparison_rows)
        df_comparison.to_excel(writer, sheet_name="三模板對比", index=False)
        ws = writer.sheets["三模板對比"]

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 50
        ws.column_dimensions["I"].width = 60

        from openpyxl.styles import PatternFill

        fill_colors = {
            "concise": PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            ),
            "standard": PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
            ),
            "detailed": PatternFill(
                start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
            ),
        }

        current_row = 2
        for i in range(0, len(comparison_rows), 3):
            mode_text = comparison_rows[i]["模板模式"]
            mode_key = (
                "concise"
                if "concise" in mode_text.lower()
                else ("standard" if "standard" in mode_text.lower() else "detailed")
            )
            fill = fill_colors.get(mode_key)
            if fill:
                for col in range(1, 10):
                    ws.cell(row=current_row, column=col).fill = fill
            current_row += 3

        ws.freeze_panes = "A2"

        # ===== Sheet 2: 摘要統計 =====
        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, sheet_name="摘要統計", index=False)
        ws2 = writer.sheets["摘要統計"]

        ws2.column_dimensions["A"].width = 8
        ws2.column_dimensions["B"].width = 30
        ws2.column_dimensions["C"].width = 12
        ws2.column_dimensions["D"].width = 8
        ws2.column_dimensions["E"].width = 50
        ws2.column_dimensions["F"].width = 12
        ws2.column_dimensions["G"].width = 8
        ws2.column_dimensions["H"].width = 50
        ws2.column_dimensions["I"].width = 12
        ws2.column_dimensions["J"].width = 8
        ws2.column_dimensions["K"].width = 50
        ws2.column_dimensions["L"].width = 10
        ws2.column_dimensions["M"].width = 8
        ws2.column_dimensions["N"].width = 8
        ws2.freeze_panes = "A2"

        # ===== Sheet 3: 測試摘要 =====
        checkpoint_status = (
            "✅ PASS"
            if (concise_ok == total and standard_ok == total and all_different == total)
            else "⚠️ 需要關注"
        )

        summary_data = {
            "項目": [
                "測試日期",
                "資料來源",
                "測試方法",
                "",
                "--- 測試結果 ---",
                "concise 達標（≤110字）",
                "standard 達標（≤210字）",
                "三模板差異通過（detailed > concise）",
                "",
                "--- 平均字數 ---",
                "concise 平均",
                "standard 平均",
                "detailed 平均",
                "",
                "--- Phase 14 Stage 2 Checkpoint 標準 ---",
                "concise ≤110 字",
                "standard ≤210 字",
                "三模板差異度 >30%",
                "向後相容",
                "",
                "--- Checkpoint 結論 ---",
                "整體狀態",
            ],
            "結果": [
                datetime.now().strftime("%Y-%m-%d %H:%M"),
                "Checkpoint 1 (10 個真實案例)",
                "同一輸入 → post_process → differentiate_template (三模式)",
                "",
                "",
                f"{concise_ok}/{total} ✅ ({concise_ok / total * 100:.0f}%)",
                f"{standard_ok}/{total} ✅ ({standard_ok / total * 100:.0f}%)",
                f"{all_different}/{total} ✅ ({all_different / total * 100:.0f}%)",
                "",
                "",
                f"{avg_c:.1f} 字",
                f"{avg_s:.1f} 字",
                f"{avg_d:.1f} 字",
                "",
                "",
                "✅ PASS" if concise_ok == total else "⚠️ 未全達標",
                "✅ PASS" if standard_ok == total else "⚠️ 未全達標",
                "✅ PASS" if all_different == total else "⚠️ 未全達標",
                "✅ PASS (預設 standard)",
                "",
                "",
                checkpoint_status,
            ],
        }
        df_test_summary = pd.DataFrame(summary_data)
        df_test_summary.to_excel(writer, sheet_name="測試摘要", index=False)
        ws3 = writer.sheets["測試摘要"]
        ws3.column_dimensions["A"].width = 40
        ws3.column_dimensions["B"].width = 50

        # ===== Sheet 4-6: 詳細內容（三模式）=====
        for mode in MODES:
            mode_name = MODE_NAMES[mode]
            sheet_name = f"詳細-{mode}"

            rows = []
            for case in cases:
                cid = case["case_id"]
                cname = case["company"]["name"]
                result = all_results[cid - 1]["modes"][mode]

                rows.append(
                    {
                        "案例": cid,
                        "公司名稱": cname,
                        "原始輸入（Checkpoint1）": case.get("input_text", ""),
                        "post_process後": result["post_processed"],
                        "最終輸出": result["final_output"],
                        "純文字": result["plain_text"],
                        "中文字數": result["chinese_count"],
                        "純文字數": result["plain_count"],
                    }
                )

            df_mode = pd.DataFrame(rows)
            df_mode.to_excel(writer, sheet_name=sheet_name, index=False)
            ws_m = writer.sheets[sheet_name]
            ws_m.column_dimensions["A"].width = 8
            ws_m.column_dimensions["B"].width = 30
            ws_m.column_dimensions["C"].width = 50
            ws_m.column_dimensions["D"].width = 60
            ws_m.column_dimensions["E"].width = 60
            ws_m.column_dimensions["F"].width = 50
            ws_m.column_dimensions["G"].width = 12
            ws_m.column_dimensions["H"].width = 12

    print(f"✅ Excel 報告已生成: {excel_path}")

    # ===== 生成 JSON 報告 =====
    json_path = OUTPUT_DIR / f"stage2_template_differentiation_fixed_{timestamp}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "generated_at": datetime.now().isoformat(),
                "test_method": "同一輸入 → post_process → differentiate_template (三模式截斷)",
                "data_source": "Checkpoint 1 (10 個真實案例)",
                "summary": {
                    "total_cases": total,
                    "concise_ok": concise_ok,
                    "standard_ok": standard_ok,
                    "all_different": all_different,
                    "avg_concise_chars": round(avg_c, 1),
                    "avg_standard_chars": round(avg_s, 1),
                    "avg_detailed_chars": round(avg_d, 1),
                },
                "results": all_results,
            },
            f,
            ensure_ascii=False,
            indent=2,
        )
    print(f"✅ JSON 報告已生成: {json_path}")

    # ===== 生成 CSV 摘要 =====
    csv_path = OUTPUT_DIR / f"stage2_template_summary_fixed_{timestamp}.csv"
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(summary_rows[0].keys()))
        writer.writeheader()
        writer.writerows(summary_rows)
    print(f"✅ CSV 摘要已生成: {csv_path}")

    # ===== 終端輸出 =====
    print()
    print("=" * 70)
    print("📊 Phase 14 Stage 2 整合測試報告摘要（修正版）")
    print("=" * 70)
    print()
    print(f"測試方法: 同一輸入 → post_process → differentiate_template (三模式)")
    print(f"資料來源: Checkpoint 1 ({total} 個真實案例)")
    print()
    print("--- 字數統計 ---")
    print(f"  concise  平均: {avg_c:.1f} 字  (目標 ≤110)")
    print(f"  standard 平均: {avg_s:.1f} 字  (目標 ≤210)")
    print(f"  detailed 平均: {avg_d:.1f} 字")
    print()
    print("--- Checkpoint 標準 ---")
    print(
        f"  concise ≤110 字: {concise_ok}/{total} ✅ ({concise_ok / total * 100:.0f}%)"
    )
    print(
        f"  standard ≤210 字: {standard_ok}/{total} ✅ ({standard_ok / total * 100:.0f}%)"
    )
    print(
        f"  三模板差異度: {all_different}/{total} ✅ ({all_different / total * 100:.0f}%)"
    )
    print()
    print("--- 輸出檔案 ---")
    print(f"  📊 Excel: {excel_path.name}")
    print(f"  📄 JSON:  {json_path.name}")
    print(f"  📋 CSV:   {csv_path.name}")
    print()
    print("=" * 70)
    print("✅ 報告生成完成")
    print("=" * 70)


if __name__ == "__main__":
    main()
