#!/usr/bin/env python3
"""
生成 Checkpoint 1 整合測試的 xlsx 報告
=======================================

參照 checkpoint1_automated_test_results.xlsx 的格式

使用方法:
  python3 generate_xlsx_report.py

"""

import json
import os
import sys
from pathlib import Path
from datetime import datetime

# 確保 src 目錄在 Python 路徑中
PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def load_checkpoint1_data():
    """載入 Checkpoint 1 測試數據"""
    data_path = (
        PROJECT_ROOT
        / "docs"
        / "test_report"
        / "v0.0.1"
        / "phase14"
        / "checkpoint1"
        / "test_data"
        / "structured_test_data.json"
    )

    with open(data_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    return data["test_cases"]


def load_integration_results():
    """載入整合測試結果"""
    report_path = (
        PROJECT_ROOT / "scripts" / "stage2" / "checkpoint1_integration_report.json"
    )

    with open(report_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    return data


def load_llm_responses():
    """載入 LLM 回應（從測試腳本的輸出）"""
    # 這個函數可以從日誌或測試輸出中提取 LLM 回應
    # 目前先返回空字典，需要手動填入
    return {}


def generate_xlsx_report():
    """生成 xlsx �告"""

    print("=" * 80)
    print("生成 Checkpoint 1 整合測試 xlsx 報告")
    print("=" * 80)

    # 載入數據
    test_cases = load_checkpoint1_data()
    integration_results = load_integration_results()

    # 創建工作簿
    wb = Workbook()

    # ===== 測試結果工作表 =====
    ws_results = wb.active
    ws_results.title = "測試結果"

    # 設置樣式
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )
    header_border = Border(bottom=Side(style="thin"))

    # 寫入標題行
    headers = [
        "案例",
        "廠商名稱 & 編號",
        "產業類型 & 代碼",
        "產品/服務",
        "公司簡介 - before",
        "公司簡介 - after",
        "Optimisation Mode",
        "Response Time",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws_results.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # 寫入測試結果
    results_by_case_id = {r["case_id"]: r for r in integration_results["results"]}

    for row_idx, case in enumerate(test_cases[:10], 2):
        case_id = case["case_id"]
        result = results_by_case_id.get(case_id, {})

        # 公司名稱 & 編號
        company_info = (
            f"{case['company']['name']}\n{case['company']['registration_id']}"
        )

        # 產業類型 & 代碼
        industry_info = f"{case['industry']['type']}\n{case['industry']['code']}"

        # 產品/服務
        product_service = case.get("product_service", "")

        # Before text
        before_text = case["before_text"]

        # After text（從測試結果中提取，移除 HTML 標籤）
        after_text = result.get("output", "")
        # 移除 <p> 和 </p> 標籤
        after_text = after_text.replace("<p>", "").replace("</p>", "")

        # Optimisation Mode（固定為 1，表示使用 Few-shot 方案）
        optimisation_mode = 1

        # Response Time
        latency = result.get("latency", 0)
        latency_str = f"{latency:.2f} s"

        # 寫入數據
        ws_results.cell(row=row_idx, column=1, value=case_id)
        ws_results.cell(row=row_idx, column=2, value=company_info)
        ws_results.cell(row=row_idx, column=3, value=industry_info)
        ws_results.cell(row=row_idx, column=4, value=product_service)
        ws_results.cell(row=row_idx, column=5, value=before_text)
        ws_results.cell(row=row_idx, column=6, value=after_text)
        ws_results.cell(row=row_idx, column=7, value=optimisation_mode)
        ws_results.cell(row=row_idx, column=8, value=latency_str)

        # 設置樣式
        for col in range(1, 9):
            cell = ws_results.cell(row=row_idx, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 調整列寬
    ws_results.column_dimensions["A"].width = 6
    ws_results.column_dimensions["B"].width = 25
    ws_results.column_dimensions["C"].width = 20
    ws_results.column_dimensions["D"].width = 20
    ws_results.column_dimensions["E"].width = 40
    ws_results.column_dimensions["F"].width = 40
    ws_results.column_dimensions["G"].width = 15
    ws_results.column_dimensions["H"].width = 12

    # ===== 測試摘要工作表 =====
    ws_summary = wb.create_sheet("測試摘要")

    # 計算統計
    total_cases = integration_results["total_cases"]
    successful_cases = integration_results["successful_cases"]
    failed_cases = integration_results["failed_cases"]
    success_rate = (successful_cases / total_cases * 100) if total_cases > 0 else 0

    results = integration_results["results"]
    avg_latency = (
        sum(r.get("latency", 0) for r in results) / len(results) if results else 0
    )
    avg_usage_rate = (
        sum(r.get("usage_rate", 0) for r in results) / len(results) if results else 0
    )

    # 寫入摘要
    summary_data = [
        ["項目", "結果"],
        ["總案例數", total_cases],
        ["成功", successful_cases],
        ["失敗", failed_cases],
        ["成功率", f"{success_rate:.1f}%"],
        ["平均資訊使用率", f"{avg_usage_rate:.1f}%"],
        ["平均響應時間", f"{avg_latency:.2f}s"],
        ["Few-shot 方案狀態", "PASSED ✅" if success_rate >= 90 else "FAILED ❌"],
        ["報告生成時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]

    for row_idx, (item, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=item)
        ws_summary.cell(row=row_idx, column=2, value=value)

        # 設置樣式
        if row_idx == 1:
            ws_summary.cell(row=row_idx, column=1).font = header_font
            ws_summary.cell(row=row_idx, column=1).fill = header_fill
            ws_summary.cell(row=row_idx, column=2).font = header_font
            ws_summary.cell(row=row_idx, column=2).fill = header_fill

    # 調整列寬
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 30

    # 保存文件
    output_path = (
        PROJECT_ROOT
        / "scripts"
        / "stage2"
        / "checkpoint1_fewshot_integration_results.xlsx"
    )
    wb.save(output_path)

    print(f"\n✅ xlsx 報告已生成: {output_path}")

    # 顯示摘要
    print(f"\n📊 測試摘要:")
    print(f"   總案例數: {total_cases}")
    print(f"   成功: {successful_cases}")
    print(f"   失敗: {failed_cases}")
    print(f"   成功率: {success_rate:.1f}%")
    print(f"   平均資訊使用率: {avg_usage_rate:.1f}%")
    print(f"   平均響應時間: {avg_latency:.2f}s")

    return output_path


if __name__ == "__main__":
    generate_xlsx_report()
